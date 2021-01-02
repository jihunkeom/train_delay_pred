from openpyxl import load_workbook
import datetime


def trainGetRealTimeInfo(csv_file):
    trains = {}
    f = open(csv_file, 'rt')

    for line in f.readlines():
        specific_time_trains = line.strip().split(",")
        if specific_time_trains[0] == '시간':
            continue
        specific_time = datetime.datetime.strptime(specific_time_trains[0], "%H:%M:%S.%f")
        specific_time = specific_time.time()
        specific_time_trains_num = (len(specific_time_trains)-1)//5

        for i in range(specific_time_trains_num):
            train_num = int(specific_time_trains[5*i+1])
            train_num = str(train_num + (1 if train_num % 2 == 1 else -1))

            if train_num in trains:
                trains[train_num].append(
                    (specific_time, specific_time_trains[5*i+2], specific_time_trains[5*i+3],
                     specific_time_trains[5*i+4], specific_time_trains[5*i+5])
                )
            else:
                trains[train_num] = [
                    (specific_time, specific_time_trains[5 * i + 2], specific_time_trains[5 * i + 3],
                     specific_time_trains[5 * i + 4], specific_time_trains[5 * i + 5])
                ]

    f.close()
    return trains


def trainGetTimeScheduleInfo(excel_filename, worksheet_name, min_row, max_row, min_col, max_col, rapid):
    # load workbook
    print('Loading timetable...')
    wb = load_workbook(excel_filename)

    schedule_trains = {}

    ws = wb[worksheet_name]

    train_nums = []
    station_names = []
    departure_stations = []
    arrival_stations = []

    """
    for col in ws.iter_cols(min_row=min_row, max_row=min_row, min_col=min_col+1, max_col=max_col):
        for cell in col:
            departure_stations.append(cell.value)
    """
    for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col, max_col=min_col):
        for cell in row:
            departure_stations.append(cell.value)

    """
    for col in ws.iter_cols(min_row=min_row+1, max_row=min_row+1, min_col=min_col+1, max_col=max_col):
        for cell in col:
            arrival_stations.append(cell.value)
    """
    for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col+1, max_col=min_col+1):
        for cell in row:
            arrival_stations.append(cell.value)

    """
    for col in ws.iter_cols(min_row=min_row+2, max_row=min_row+2, min_col=min_col+1, max_col=max_col):
        for cell in col:
            schedule_trains[cell.value[1:]] = []
            train_nums.append(cell.value[1:])
    """
    for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col+2, max_col=min_col+2):
        for cell in row:
            schedule_trains[cell.value[1:]] = []
            train_nums.append(cell.value[1:])

    """
    for row in ws.iter_rows(min_row=min_row+3, max_row=max_row, min_col=min_col, max_col=min_col):
        for cell in row:
            station_names.append(cell.value)
    """
    for col in ws.iter_cols(min_row=min_row, max_row=min_row, min_col=min_col+(4 if rapid else 3), max_col=max_col):
        for cell in col:
            station_names.append(cell.value)

    """
    for col, num in zip(ws.iter_cols(min_row=min_row+3, max_row=max_row, min_col=min_col+1, max_col=max_col),
                        train_nums):
        flag_append = False
        arrival_and_depart = []
        for cell, stn in zip(col, station_names):
            if not flag_append:
                arrival_and_depart.append(stn)
    
            if isinstance(cell.value, str) and cell.value.strip() == '':
                arrival_and_depart.append(None)
            else:
                arrival_and_depart.append(cell.value)
    
            if flag_append:
                schedule_trains[num].append(arrival_and_depart)
                arrival_and_depart = []
            flag_append = not flag_append
    """
    for row, num in zip(ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col + (4 if rapid else 3),
                                     max_col=max_col), train_nums):
        is_departure = True
        for cell, stn in zip(row, station_names):
            if isinstance(cell.value, str):
                tm = cell.value.strip()
                if tm == "":
                    tm = None
                else:
                    tm = datetime.datetime.strptime(tm, "%H:%M:%S").time()
            else:
                tm = cell.value

            if is_departure:
                schedule_trains[num].append((stn, None, tm))
            else:
                schedule_trains[num].append((stn, tm, tm))

            if is_departure and tm is not None:
                is_departure = False

    return schedule_trains, departure_stations, arrival_stations


def total_seconds(tm):
    return (tm.hour * 60 + tm.minute) * 60 + tm.second


stn_convert_table = {
    '문산': '문산',
    '파주': '파주',
    '월롱': '월롱',
    '금촌': '금촌',
    '금릉': '금릉',
    '운정': '운정',
    '야당': '야당',
    '탄현': '탄현',
    '일산': '일산',
    '풍산': '풍산',
    '백마': '백마',
    '곡산': '곡산',
    '대곡': '대곡',
    '능곡': '능곡',
    '행신': '행신',
    '강매': '강매',
    '화전': '화전',
    '수색': '수색',
    '디엠시': '디지털미디어시티',
    '가좌': '가좌',
    '홍대입': '홍대입구',
    '서강대': '서강대',
    '공덕': '공덕',
    '효창공': '효창공원앞',
    '용산': '용산',
    '이촌': '이촌',
    '서빙고': '서빙고',
    '한남': '한남',
    '옥수': '옥수',
    '응봉': '응봉',
    '왕십리': '왕십리',
    '청량리': '청량리',
    '회기': '회기',
    '중랑': '중랑',
    '상봉': '상봉',
    '망우': '망우',
    '양원': '양원',
    '1양원': '양원',
    '2양원': '양원',
    '구리': '구리',
    '도농': '도농',
    '1양정': '양정',
    '2양정': '양정',
    '덕소': '덕소',
    '도심': '도심',
    '팔당': '팔당',
    '운길산': '운길산',
    '양수': '양수',
    '신원': '신원',
    '국수': '국수',
    '아신': '아신',
    '오빈': '오빈',
    '양평': '양평',
    '원덕': '원덕',
    '용문': '용문',
    '지평': '지평'
}

stn_list = [
    '지평', '용문', '원덕', '양평', '오빈', '아신', '국수', '신원',
    '양수', '운길산', '팔당', '도심', '덕소', '양정', '도농', '구리', '양원', '망우', '상봉',
    '중랑', '회기', '청량리', '왕십리', '응봉', '옥수', '한남', '서빙고', '이촌', '용산', '효창공원앞',
    '공덕', '서강대', '홍대입구', '가좌', '디지털미디어시티', '수색', '화전', '강매', '행신', '능곡', '대곡',
    '곡산', '백마', '풍산', '일산', '탄현', '야당', '운정', '금릉', '금촌', '월롱', '파주', '문산'
]

if __name__ == "__main__":
    date = '2020-12-12'
    # week = ''
    week = '휴 '

    trains = trainGetRealTimeInfo(f'../SubwayAPICrawling/{date}.csv')
    schedule_trains, up_depart, up_arrival = trainGetTimeScheduleInfo('경의중앙선(2020년 11월27일)출발시각.xlsx',
                                                                      f'경의중앙선({week}상)',
                                                                      2, 93 if week == '' else 76, 2,
                                                                      58 if week == '' else 57, week == '')
    schedule_trains_down, down_depart, down_arrival = trainGetTimeScheduleInfo('경의중앙선(2020년 11월27일)출발시각.xlsx',
                                                                               f'경의중앙선({week}하)',
                                                                               2, 94 if week == '' else 77, 2,
                                                                               58 if week == '' else 57, week == '')
    schedule_trains.update(schedule_trains_down)

    # print(trains)
    # print(schedule_trains)

    f_up = open(f'{date}-up-delay.csv', 'wt')
    f_down = open(f'{date}-down-delay.csv', 'wt')

    f_up.write('시발,종착,열차번호,' + ','.join(stn_list) + '\n')
    f_down.write('시발,종착,열차번호,' + ','.join(list(reversed(stn_list))) + '\n')

    for train_idx, train_num in enumerate(schedule_trains.keys()):
        # print(f"{train_num} -> {train_num in trains}")
        if train_num in trains:
            if int(train_num) % 2 == 0:
                f_up.write(f"{up_depart[train_idx]},{up_arrival[train_idx]},K{train_num}")
            else:
                f_down.write(f"{down_depart[train_idx-len(up_depart)]},{down_arrival[train_idx-len(up_arrival)]},K{train_num}")

            ini_moving = trains[train_num][0]
            ini_current_pos = ini_moving[1]

            sc_pos = 0
            delay_list = {}
            approaching_only_list = []
            approaching_list = []
            for moving in trains[train_num]:
                sc = schedule_trains[train_num]
                if sc_pos >= len(sc):
                    break
                tm, current_pos, _, approaching, express = moving
                sc_station, sc_arrival, _ = sc[sc_pos]
                while True:
                    if int(train_num) % 2 == 0 and stn_list.index(current_pos) <= stn_list.index(stn_convert_table[sc_station]):
                        break
                    if int(train_num) % 2 == 1 and stn_list.index(current_pos) >= stn_list.index(stn_convert_table[sc_station]):
                        break
                    sc_pos += 1
                    sc_station, sc_arrival, _ = sc[sc_pos]

                if sc_arrival is not None:
                    if approaching == '1' and current_pos not in delay_list:
                        delay_list[current_pos] = total_seconds(tm) - total_seconds(sc_arrival)
                        if delay_list[current_pos] <= -80000:
                            delay_list[current_pos] += 86400
                        # print(f"{current_pos} - {sc_station} : {delay_list[current_pos]}")
                        if current_pos in approaching_only_list:
                            approaching_only_list.remove(current_pos)
                    if approaching == '0' and current_pos not in approaching_only_list and current_pos not in delay_list:
                        approaching_only_list.append(current_pos)
                    if approaching == '0':
                        approaching_list.append(current_pos)

            # calculate average(middle) delay time between 2 stations' approaching time in approaching_only_list
            for stn in approaching_only_list:
                idx = stn_list.index(stn)
                if int(train_num) % 2 == 0 and idx+1 >= len(stn_list):
                    continue
                if int(train_num) % 2 == 1 and idx-1 < 0:
                    continue
                next_stn = stn_list[idx+1 if int(train_num) % 2 == 0 else idx-1]

                if next_stn not in approaching_list:
                    continue

                # print(stn, next_stn)

                sc_current_stn_idx = 0
                sc_current_stn_arrival = None
                # find train's position for searching
                for i in range(len(schedule_trains[train_num])):
                    sc_station, sc_arrival, _ = schedule_trains[train_num][i]
                    if stn_convert_table[sc_station] == stn:
                        sc_current_stn_idx = i
                        sc_current_stn_arrival = sc_arrival
                        # print(f"{train_num}: start position = {current_stn_pos}, station = {station}")
                        break

                current_stn_idx = 0
                for i in range(len(trains[train_num])):
                    tm, current_pos, _, approaching, express = trains[train_num][i]
                    if current_pos == stn and approaching == '0':
                        current_stn_idx = i
                        break

                current_stn_approach_time = total_seconds(trains[train_num][current_stn_idx][0])

                next_stn_idx = 0
                for i in range(len(trains[train_num])):
                    tm, current_pos, _, approaching, express = trains[train_num][i]
                    if current_pos == next_stn and approaching == '0':
                        next_stn_idx = i
                        break

                next_stn_approach_time = total_seconds(trains[train_num][next_stn_idx][0])

                delay_list[stn] = int((current_stn_approach_time + next_stn_approach_time) / 2
                                      - total_seconds(sc_current_stn_arrival))

            # print(delay_list)
            # print(approaching_list)
            # print(approaching_only_list)

            # write to file
            if int(train_num) % 2 == 0:
                for stn in stn_list:
                    if stn in delay_list:
                        f_up.write(',' + str(delay_list[stn]))
                    else:
                        f_up.write(',')
                f_up.write('\n')
            else:
                for stn in reversed(stn_list):
                    if stn in delay_list:
                        f_down.write(',' + str(delay_list[stn]))
                    else:
                        f_down.write(',')
                f_down.write('\n')

    f_up.close()
    f_down.close()
